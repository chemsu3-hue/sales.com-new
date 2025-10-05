# sales_app_streamlit.py — cute UI + tiles autofill + escribe "Nombre del Artículo" en columna D
import streamlit as st
import pandas as pd 
import os, unicodedata, re
from datetime import date, datetime
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

# ==============================
# Config
# ==============================
EXCEL_FILE    = "mimamuni sales datta+.xlsx"
TARGET_SHEET  = "Sheet1"
CAT_SHEET     = "Catalogo"

# Forzar "Nombre del Artículo" a la columna D (índice 4)
FORCE_ARTICULO_COLUMN_LETTER = "D"
FORCE_ARTICULO_COL = column_index_from_string(FORCE_ARTICULO_COLUMN_LETTER)

EXPECTED = ["Fecha","Cantidad","Nombre del Artículo","Método de Pago","Precio Unitario","Venta Total","Comentarios"]

DEFAULT_CATALOG = [
    {"Artículo":"bolsa","Precio":120.0},
    {"Artículo":"jeans","Precio":50.0},
    {"Artículo":"t-shirt","Precio":25.0},
    {"Artículo":"jacket","Precio":120.0},
    {"Artículo":"cinturón","Precio":20.0},
]

# ==============================
# Cute styling 🎀
# ==============================
st.set_page_config(page_title="Ventas - Tienda MIMAMUNI", page_icon="🛍️", layout="wide")
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Poppins:wght@500;700&display=swap');

:root {
  --brand: #7c5cff;
  --brand-2: #ff7ac8;
  --card-bg: #ffffff;
  --muted: #6b7280;
}

/* Page padding */
.block-container{padding-top:1.2rem}

/* Big title with gradient */
h1 {
  font-family: 'Poppins', system-ui, -apple-system, Segoe UI, Roboto, Arial, sans-serif;
  font-weight: 700 !important;
  background: linear-gradient(90deg, var(--brand), var(--brand-2));
  -webkit-background-clip: text; background-clip: text; color: transparent !important;
  letter-spacing: 0.3px;
}

/* Section subtitles */
h2, .nice-title {
  font-family: 'Poppins', system-ui, sans-serif;
  font-weight: 600;
  color: #1f2937;
  margin: 0.25rem 0 0.6rem 0;
}

/* Make cards/blocks feel soft */
div[data-testid="stVerticalBlock"] > div:has(> div > div[data-testid="stFileUploader"]) {
  background: var(--card-bg);
  border: 1px solid #eef2ff;
  border-radius: 16px;
  padding: 16px 16px 6px 16px;
  box-shadow: 0 2px 12px rgba(124,92,255,0.08);
}

/* Data editor soft corners */
div[data-testid="stDataFrame"] {
  border-radius: 14px; overflow: hidden; border: 1px solid #eef2ff;
}

/* Buttons general */
button[kind="primary"] {
  border-radius: 12px !important;
  padding: 10px 14px !important;
  font-weight: 700 !important;
}
button[kind="secondary"] {
  border-radius: 12px !important;
  padding: 10px 12px !important;
  font-weight: 600 !important;
}

/* Tile buttons (we use secondary kind for tiles) */
.tile-grid button[kind="secondary"] {
  min-height: 64px;
  font-size: 0.95rem;
  white-space: pre-line !important;
  background: linear-gradient(135deg, #ffffff, #f7f3ff);
  border: 1px solid #e9e5ff;
  box-shadow: 0 2px 10px rgba(124,92,255,0.10);
  transition: transform .06s ease, box-shadow .2s ease;
}
.tile-grid button[kind="secondary"]:hover {
  transform: translateY(-1px);
  box-shadow: 0 6px 18px rgba(124,92,255,0.18);
}

/* Small helper text */
.small { color: var(--muted); font-size: 0.9rem; }
</style>
""", unsafe_allow_html=True)

st.title("🛍️ Ventas — Tienda mimamuni")

# ==============================
# Helpers
# ==============================
def ensure_excel_exists() -> bool:
    return os.path.exists(EXCEL_FILE)

def open_wb():
    if not ensure_excel_exists():
        raise FileNotFoundError("Excel no encontrado. Sube tu archivo arriba.")
    return load_workbook(EXCEL_FILE)

def write_sheet_replace(df: pd.DataFrame, sheet_name: str):
    """Reemplaza/crea una hoja con el contenido de df."""
    from openpyxl.utils.dataframe import dataframe_to_rows
    wb = open_wb()
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    ws = wb.create_sheet(sheet_name)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    wb.save(EXCEL_FILE)

def canon(s: str) -> str:
    """Normaliza: sin acentos, espacios colapsados, minúsculas."""
    if s is None: return ""
    s = str(s).replace("\u00A0", " ")
    s = unicodedata.normalize("NFKD", s).encode("ascii","ignore").decode("ascii")
    s = re.sub(r"\s+", " ", s).strip().lower()
    return s

HEADER_SYNONYMS = {
    "fecha":"Fecha",
    "cantidad":"Cantidad",
    "nombre del articulo":"Nombre del Artículo",
    "nombre del artículo":"Nombre del Artículo",
    "articulo":"Nombre del Artículo",
    "artículo":"Nombre del Artículo",
    "producto":"Nombre del Artículo",
    "descripcion":"Nombre del Artículo",
    "descripción":"Nombre del Artículo",
    "metodo de pago":"Método de Pago",
    "metodo pago":"Método de Pago",
    "precio unitario":"Precio Unitario",
    "venta total":"Venta Total",
    "comentarios":"Comentarios",
    "comentario":"Comentarios",
}

def detect_headers(ws, max_cols: int = 60):
    """Devuelve (fila_cabecera, lista_cabeceras_raw) detectando por Fecha/Cantidad/Nombre Artículo."""
    max_rows = min(ws.max_row, 300)
    def row_vals(r: int):
        return [ws.cell(r, c).value for c in range(1, max_cols+1)]
    for r in range(1, max_rows+1):
        vals = row_vals(r)
        can = [canon(v) for v in vals]
        if {"fecha","cantidad","nombre del articulo"} <= set(can):
            return r, vals
    return None, []

def build_col_map(headers_raw):
    """Mapa estándar -> índice columna. Luego forzamos Artículo a D."""
    cmap = {}
    for idx, h in enumerate(headers_raw, start=1):
        std = HEADER_SYNONYMS.get(canon(h))
        if std in EXPECTED:
            cmap[std] = idx
    return cmap

def next_row_by_fecha(ws, header_row: int, fecha_col: int):
    """Primera fila vacía bajo la última con Fecha."""
    r = header_row + 1
    last = header_row
    while r <= ws.max_row:
        if ws.cell(r, fecha_col).value not in (None, ""):
            last = r; r += 1
        else:
            break
    return last + 1

def append_sale_to_sheet(row: dict) -> dict:
    """Escribe la venta en TARGET_SHEET. 'Nombre del Artículo' SIEMPRE en columna D."""
    wb = open_wb()
    if TARGET_SHEET not in wb.sheetnames:
        raise ValueError(f"No se encontró la hoja '{TARGET_SHEET}'.")
    ws = wb[TARGET_SHEET]

    hr, headers_raw = detect_headers(ws)
    if not hr:
        raise RuntimeError("No se detectó la fila de cabeceras (Fecha/Cantidad/Nombre...).")
    cmap = build_col_map(headers_raw)

    # Forzar Artículo -> columna D
    cmap["Nombre del Artículo"] = FORCE_ARTICULO_COL

    if "Fecha" not in cmap:
        raise RuntimeError("No se encontró la columna 'Fecha' en la tabla.")
    write_row = next_row_by_fecha(ws, hr, cmap["Fecha"])

    # Venta Total auto si falta
    if "Venta Total" in cmap and (row.get("Venta Total") in (None, "")):
        try: row["Venta Total"] = float(row.get("Cantidad",0)) * float(row.get("Precio Unitario",0))
        except Exception: row["Venta Total"] = None

    # Escribir solo columnas conocidas (con D forzada para Artículo)
    for std, col in cmap.items():
        val = row.get(std, None)
        if std == "Nombre del Artículo" and val is not None:
            val = str(val)
        if std == "Fecha" and isinstance(val, (date, datetime)):
            ws.cell(write_row, col).value = datetime.combine(val, datetime.min.time())
        else:
            ws.cell(write_row, col).value = val

    wb.save(EXCEL_FILE)
    return {"header_row": hr, "written_row": write_row, "columns_used": cmap}

@st.cache_data(show_spinner=False)
def load_catalog_df() -> pd.DataFrame:
    if not ensure_excel_exists():
        return pd.DataFrame(DEFAULT_CATALOG, columns=["Artículo","Precio"])
    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name=CAT_SHEET)
        df = df.rename(columns={"Articulo":"Artículo","precio":"Precio"})
        df["Artículo"] = df["Artículo"].astype(str)
        df["Precio"]   = pd.to_numeric(df["Precio"], errors="coerce").fillna(0.0)
        return df[["Artículo","Precio"]]
    except Exception:
        df = pd.DataFrame(DEFAULT_CATALOG, columns=["Artículo","Precio"])
        try: write_sheet_replace(df, CAT_SHEET)
        except Exception: pass
        return df

def save_catalog_df(df: pd.DataFrame):
    clean = df.copy()
    clean["Artículo"] = clean["Artículo"].astype(str).str.strip()
    clean["Precio"]   = pd.to_numeric(clean["Precio"], errors="coerce").fillna(0.0)
    clean = clean[clean["Artículo"]!=""].drop_duplicates(subset=["Artículo"], keep="last")
    write_sheet_replace(clean[["Artículo","Precio"]], CAT_SHEET)

# ==============================
# 1) Excel subir/descargar
# ==============================
st.markdown("### 📂 Archivo Excel")
up = st.file_uploader("Sube tu Excel (.xlsx). Se guardará en la tabla de Sheet1 (VENTAS DIARIAS).", type=["xlsx"])
if up is not None:
    with open(EXCEL_FILE, "wb") as f: f.write(up.getbuffer())
    st.success("Excel guardado.")
    st.cache_data.clear()

if ensure_excel_exists():
    with open(EXCEL_FILE, "rb") as f:
        st.download_button("⬇️ Descargar Excel actualizado", f, file_name=EXCEL_FILE,
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
else:
    st.info("Aún no has subido el archivo.")

# ==============================
# 2) Catálogo CRUD (persistente)
# ==============================
st.divider()
st.markdown("### 🗂️ Catálogo de artículos")
st.markdown("<span class='small'>Añade, edita o borra. Se guarda en la hoja ‘Catalogo’ del mismo Excel.</span>", unsafe_allow_html=True)

cat_df = load_catalog_df()
if "Eliminar" not in cat_df.columns:
    cat_df["Eliminar"] = False

edited = st.data_editor(
    cat_df, num_rows="dynamic", hide_index=True, use_container_width=True,
    column_config={
        "Artículo": st.column_config.TextColumn(required=True),
        "Precio":   st.column_config.NumberColumn(min_value=0.0, step=1.0, format="%.2f"),
        "Eliminar": st.column_config.CheckboxColumn(help="Marca para borrar esta fila"),
    },
    key="catalog_editor"
)

c1, c2 = st.columns([1,1])
with c1:
    if st.button("💾 Guardar catálogo", use_container_width=True, disabled=not ensure_excel_exists()):
        to_save = edited.copy()
        if "Eliminar" in to_save:
            to_save = to_save[to_save["Eliminar"]==False].drop(columns=["Eliminar"])
        save_catalog_df(to_save)
        st.success("Catálogo guardado en ‘Catalogo’.")
        st.cache_data.clear(); st.rerun()
with c2:
    if st.button("↩️ Deshacer cambios", use_container_width=True):
        st.cache_data.clear(); st.rerun()

# ==============================
# 3) Tiles (ELIGE UN ARTÍCULO) — más pequeños y bonitos
# ==============================
st.divider()
st.markdown("### 🧱 Elige un artículo")
st.markdown("<span class='small'>Clic en un tile → rellena automáticamente el formulario de venta.</span>", unsafe_allow_html=True)

tiles = load_catalog_df().sort_values("Artículo").reset_index(drop=True)

# Estado para el formulario
if "nombre_input" not in st.session_state: st.session_state.nombre_input = ""
if "precio_sel"   not in st.session_state: st.session_state.precio_sel   = 0.0

# Búsqueda
search = st.text_input("🔎 Buscar", placeholder="Escribe para filtrar…", label_visibility="collapsed")
if search:
    tiles = tiles[tiles["Artículo"].str.contains(search, case=False, na=False)]

# Grid de tiles (más pequeño): 5 por fila
per_row = 5
items = list(tiles.itertuples(index=False, name=None))  # [(Artículo, Precio)]
st.markdown('<div class="tile-grid">', unsafe_allow_html=True)
for i in range(0, len(items), per_row):
    cols = st.columns(per_row)
    for col, (name, price) in zip(cols, items[i:i+per_row]):
        with col:
            if st.button(f"{name}\n${float(price):.2f}", key=f"tile_{name}", use_container_width=True):
                # ✅ tile click -> rellena el formulario
                st.session_state.nombre_input = name
                st.session_state.precio_sel   = float(price)
st.markdown('</div>', unsafe_allow_html=True)

# Añadir rápido (mini formulario)
with st.expander("➕ Añadir artículo rápido", expanded=False):
    q1, q2, q3 = st.columns([2,1,1])
    with q1: qa_name = st.text_input("Nombre", key="qa_name")
    with q2: qa_price = st.number_input("Precio", min_value=0.0, step=1.0, value=0.0, format="%.2f", key="qa_price")
    with q3:
        if st.button("Guardar", use_container_width=True, key="qa_btn", disabled=not ensure_excel_exists()):
            dfc = load_catalog_df()
            mask = dfc["Artículo"].str.lower().eq((qa_name or "").strip().lower())
            if mask.any():
                dfc.loc[mask, "Precio"] = float(qa_price)
            else:
                dfc = pd.concat([dfc, pd.DataFrame([{"Artículo": (qa_name or "").strip(), "Precio": float(qa_price)}])], ignore_index=True)
            save_catalog_df(dfc)
            st.success(f"Guardado: {qa_name} → {float(qa_price):.2f}")
            st.cache_data.clear(); st.rerun()

# ==============================
# 4) Formulario de venta (cute) — escribe Artículo en columna D
# ==============================
st.divider()
st.markdown("### 🧾 Guardar venta")
st.markdown("<span class='small'>“Nombre del Artículo” se escribe siempre en la <b>columna D</b> del Excel.</span>", unsafe_allow_html=True)

left, right = st.columns(2)
with left:
    fecha    = st.date_input("📅 Fecha", value=date.today())
    cantidad = st.number_input("🔢 Cantidad", min_value=1, step=1, value=1)
    # ✅ este input se rellena con los tiles
    articulo = st.text_input("🏷️ Nombre del Artículo", key="nombre_input", placeholder="p. ej., jeans slim")
with right:
    metodo      = st.radio("💳 Método de Pago", ["E","T"], horizontal=True, help="E=Efectivo, T=Tarjeta")
    precio_unit = st.number_input("💵 Precio Unitario", min_value=0.0, step=1.0, value=float(st.session_state.precio_sel), format="%.2f")
    venta_total = st.number_input("🧮 Venta Total (auto)", min_value=0.0, step=1.0, value=float(cantidad)*float(precio_unit), format="%.2f")
comentarios = st.text_area("📝 Comentarios (opcional)")

disabled = (not ensure_excel_exists()) or (not articulo) or (precio_unit <= 0)

if st.button("💾 Guardar venta en Excel", type="primary", use_container_width=True, disabled=disabled):
    try:
        info = append_sale_to_sheet({
            "Fecha": fecha,
            "Cantidad": int(cantidad),
            "Nombre del Artículo": (articulo or "").strip(),
            "Método de Pago": metodo,
            "Precio Unitario": float(precio_unit),
            "Venta Total": float(venta_total),
            "Comentarios": (comentarios or "").strip() or None,
        })
        st.success("✅ Venta guardada (Nombre del Artículo → columna D).")
        st.info(f"Cabeceras en fila {info['header_row']} → escrita la fila **{info['written_row']}**.")
        st.balloons()
        # limpiar selección para el siguiente registro
        st.session_state.nombre_input = ""
        st.session_state.precio_sel = 0.0
        st.cache_data.clear()
    except Exception as e:
        st.error(f"No se pudo escribir: {e}")

